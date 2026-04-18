from __future__ import annotations

import csv
import datetime as dt
import io
import uuid
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.auth import get_current_admin
from app.db import get_db
from app.models import Campaign, CampaignRecipient, Customer, Event
from app.schemas import (
    CustomerCreate,
    CustomerImportResult,
    CustomerOut,
    CustomerTimelineOut,
    CustomerUpdate,
    ImportRowError,
    OkOut,
    TimelineItem,
)

router = APIRouter(prefix="/admin/customers", dependencies=[Depends(get_current_admin)])


def _customer_to_out(c: Customer) -> CustomerOut:
    return CustomerOut(
        id=str(c.id),
        email=c.email,
        name=c.name,
        company=c.company,
        phone=c.phone,
        role=c.role,
        tags=list(c.tags or []),
        source=c.source,
        notes=c.notes,
        consent_marketing=c.consent_marketing,
        consent_source=c.consent_source,
        consent_at=c.consent_at,
        created_at=c.created_at,
        updated_at=c.updated_at,
    )


@router.get("", response_model=list[CustomerOut])
def list_customers(
    search: str | None = Query(default=None),
    tag: str | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
):
    q = select(Customer)
    if search and search.strip():
        like = f"%{search.strip()}%"
        q = q.where(
            or_(
                Customer.email.ilike(like),
                Customer.name.ilike(like),
                Customer.company.ilike(like),
            )
        )
    if tag:
        q = q.where(Customer.tags.any(tag))
    q = q.order_by(Customer.created_at.desc()).offset(offset).limit(limit)
    rows = db.execute(q).scalars().all()
    return [_customer_to_out(c) for c in rows]


@router.post("", response_model=CustomerOut)
def create_customer(body: CustomerCreate, db: Session = Depends(get_db)):
    email = str(body.email).strip().lower()
    if db.scalar(select(Customer).where(Customer.email == email)) is not None:
        raise HTTPException(status_code=400, detail="Customer with this email already exists")
    c = Customer(
        id=uuid.uuid4(),
        email=email,
        name=body.name,
        company=body.company,
        phone=body.phone,
        role=body.role,
        tags=body.tags or [],
        source=body.source,
        notes=body.notes,
        consent_marketing=body.consent_marketing,
        consent_source=body.consent_source,
        consent_at=dt.datetime.now(dt.timezone.utc) if body.consent_marketing else None,
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return _customer_to_out(c)


@router.get("/{customer_id}", response_model=CustomerOut)
def get_customer(customer_id: str, db: Session = Depends(get_db)):
    c = db.get(Customer, customer_id)
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")
    return _customer_to_out(c)


@router.patch("/{customer_id}", response_model=CustomerOut)
def update_customer(customer_id: str, body: CustomerUpdate, db: Session = Depends(get_db)):
    c = db.get(Customer, customer_id)
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")
    if body.email is not None:
        new_email = str(body.email).strip().lower()
        other = db.scalar(
            select(Customer).where(Customer.email == new_email, Customer.id != c.id)
        )
        if other is not None:
            raise HTTPException(status_code=400, detail="Email already in use")
        c.email = new_email
    for field in ("name", "company", "phone", "role", "source", "notes", "consent_source"):
        v = getattr(body, field)
        if v is not None:
            setattr(c, field, v)
    if body.tags is not None:
        c.tags = body.tags
    if body.consent_marketing is not None:
        previous = c.consent_marketing
        c.consent_marketing = body.consent_marketing
        if body.consent_marketing and not previous:
            c.consent_at = dt.datetime.now(dt.timezone.utc)
    db.commit()
    db.refresh(c)
    return _customer_to_out(c)


@router.delete("/{customer_id}", response_model=OkOut)
def delete_customer(customer_id: str, db: Session = Depends(get_db)):
    c = db.get(Customer, customer_id)
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")
    db.delete(c)
    db.commit()
    return OkOut()


@router.get("/{customer_id}/timeline", response_model=CustomerTimelineOut)
def customer_timeline(customer_id: str, db: Session = Depends(get_db)):
    c = db.get(Customer, customer_id)
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")

    items: list[TimelineItem] = []

    events = (
        db.execute(
            select(Event)
            .where(Event.customer_id == c.id)
            .order_by(Event.occurred_at.desc())
            .limit(200)
        )
        .scalars()
        .all()
    )
    for ev in events:
        items.append(
            TimelineItem(
                kind=f"event:{ev.type}",
                occurred_at=ev.occurred_at,
                label=ev.type,
                data={"url": ev.url, **(ev.payload or {})},
            )
        )

    recipients = (
        db.execute(
            select(CampaignRecipient, Campaign)
            .join(Campaign, Campaign.id == CampaignRecipient.campaign_id)
            .where(CampaignRecipient.email == c.email)
            .order_by(CampaignRecipient.created_at.desc())
            .limit(200)
        ).all()
    )
    for recip, camp in recipients:
        # Emit a single row per notable campaign interaction, newest interaction first.
        for label, when in [
            ("campaign:unsubscribed", recip.unsubscribed_at),
            ("campaign:complained", recip.complained_at),
            ("campaign:bounced", recip.bounced_at),
            ("campaign:clicked", recip.clicked_at),
            ("campaign:opened", recip.opened_at),
            ("campaign:delivered", recip.delivered_at),
            ("campaign:sent", recip.sent_at),
        ]:
            if when is not None:
                items.append(
                    TimelineItem(
                        kind=label,
                        occurred_at=when,
                        label=f"{camp.name} — {label.split(':', 1)[1]}",
                        data={"campaign_id": str(camp.id), "status": recip.status},
                    )
                )
                break
        else:
            items.append(
                TimelineItem(
                    kind="campaign:queued",
                    occurred_at=recip.created_at,
                    label=f"{camp.name} — queued",
                    data={"campaign_id": str(camp.id), "status": recip.status},
                )
            )

    items.sort(key=lambda x: x.occurred_at, reverse=True)
    return CustomerTimelineOut(customer=_customer_to_out(c), items=items[:500])


# --------------------------------------------------------------------------
# Import
# --------------------------------------------------------------------------


EXPECTED_HEADERS = (
    "email",
    "name",
    "company",
    "phone",
    "role",
    "tags",
    "source",
    "notes",
    "consent_marketing",
)


def _normalize_header(h: str | None) -> str:
    if h is None:
        return ""
    return h.strip().lower().replace(" ", "_")


def _parse_csv(raw: bytes) -> list[dict[str, str]]:
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV has no header row")
    fieldmap = {_normalize_header(f): f for f in reader.fieldnames if f}
    out: list[dict[str, str]] = []
    for line in reader:
        row: dict[str, str] = {}
        for key in EXPECTED_HEADERS:
            src = fieldmap.get(key)
            row[key] = (line.get(src, "") if src else "") or ""
        out.append(row)
    return out


def _parse_xlsx(raw: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel file is empty")
    headers = [_normalize_header(str(c) if c is not None else "") for c in header_row]
    idx = {h: i for i, h in enumerate(headers) if h}

    def cell(r: tuple[Any, ...], key: str) -> str:
        i = idx.get(key)
        if i is None or i >= len(r):
            return ""
        v = r[i]
        if v is None:
            return ""
        return str(v).strip()

    out: list[dict[str, str]] = []
    for r in rows_iter:
        if all((cell(r, k) == "" for k in EXPECTED_HEADERS)):
            continue
        out.append({k: cell(r, k) for k in EXPECTED_HEADERS})
    return out


def _parse_tags(raw: str) -> list[str]:
    if not raw:
        return []
    return [t.strip() for t in raw.split(",") if t.strip()]


def _parse_bool(raw: str) -> bool:
    return raw.strip().lower() in ("1", "true", "yes", "y", "t")


@router.post("/import", response_model=CustomerImportResult)
async def import_customers(
    file: UploadFile = File(...),
    dry_run: bool = Query(default=False),
    db: Session = Depends(get_db),
):
    raw = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".csv"):
        try:
            rows = _parse_csv(raw)
        except UnicodeDecodeError as e:
            raise HTTPException(status_code=400, detail=f"CSV must be UTF-8: {e}") from e
    elif name.endswith(".xlsx"):
        rows = _parse_xlsx(raw)
    else:
        raise HTTPException(status_code=400, detail="Upload a .csv or .xlsx file")

    result = CustomerImportResult()
    now = dt.datetime.now(dt.timezone.utc)
    for i, row in enumerate(rows, start=2):
        email_raw = (row.get("email") or "").strip().lower()
        if not email_raw or "@" not in email_raw:
            result.errors.append(ImportRowError(row=i, message="valid email is required"))
            continue
        tags = _parse_tags(row.get("tags") or "")
        consent = _parse_bool(row.get("consent_marketing") or "")

        existing = db.scalar(select(Customer).where(Customer.email == email_raw))
        if dry_run:
            if existing:
                result.updated += 1
            else:
                result.created += 1
            continue

        if existing:
            existing.name = (row.get("name") or "").strip() or existing.name
            existing.company = (row.get("company") or "").strip() or existing.company
            existing.phone = (row.get("phone") or "").strip() or existing.phone
            existing.role = (row.get("role") or "").strip() or existing.role
            existing.source = (row.get("source") or "").strip() or existing.source
            existing.notes = (row.get("notes") or "").strip() or existing.notes
            if tags:
                existing.tags = sorted(set((existing.tags or []) + tags))
            if consent and not existing.consent_marketing:
                existing.consent_marketing = True
                existing.consent_source = "import"
                existing.consent_at = now
            db.commit()
            result.updated += 1
        else:
            c = Customer(
                id=uuid.uuid4(),
                email=email_raw,
                name=(row.get("name") or "").strip() or None,
                company=(row.get("company") or "").strip() or None,
                phone=(row.get("phone") or "").strip() or None,
                role=(row.get("role") or "").strip() or None,
                tags=tags,
                source=(row.get("source") or "import").strip() or "import",
                notes=(row.get("notes") or "").strip() or None,
                consent_marketing=consent,
                consent_source="import" if consent else None,
                consent_at=now if consent else None,
            )
            db.add(c)
            db.commit()
            result.created += 1

    return result
