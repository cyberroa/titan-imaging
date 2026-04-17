from __future__ import annotations

import csv
import io
import uuid
from decimal import Decimal
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.auth import get_current_admin
from app.db import get_db
from app.inventory_alerts import notify_subscribers_if_became_available
from app.models import Category, Part
from app.part_utils import is_part_available, refresh_part_search_vector
from app.schemas import (
    CategoryCreate,
    CategoryOut,
    CategoryUpdate,
    ImportResult,
    ImportRowError,
    OkOut,
    OutreachSendIn,
    OutreachSendOut,
    PartCreate,
    PartOut,
    PartUpdate,
)

router = APIRouter(prefix="/admin", dependencies=[Depends(get_current_admin)])


def _slugify(s: str) -> str:
    return "-".join(s.strip().lower().replace("_", "-").split())


@router.get("/categories", response_model=list[CategoryOut])
def admin_list_categories(db: Session = Depends(get_db)):
    rows = db.execute(select(Category).order_by(Category.name.asc())).scalars().all()
    return [CategoryOut(id=str(c.id), name=c.name, slug=c.slug) for c in rows]


@router.post("/categories", response_model=CategoryOut)
def admin_create_category(body: CategoryCreate, db: Session = Depends(get_db)):
    slug = body.slug.strip() if body.slug else _slugify(body.name)
    if db.scalar(select(Category).where(Category.slug == slug)) is not None:
        raise HTTPException(status_code=400, detail="Category slug already exists")
    c = Category(id=uuid.uuid4(), name=body.name.strip(), slug=slug)
    db.add(c)
    db.commit()
    db.refresh(c)
    return CategoryOut(id=str(c.id), name=c.name, slug=c.slug)


@router.patch("/categories/{category_id}", response_model=CategoryOut)
def admin_update_category(
    category_id: str, body: CategoryUpdate, db: Session = Depends(get_db)
):
    c = db.get(Category, category_id)
    if not c:
        raise HTTPException(status_code=404, detail="Category not found")
    if body.name is not None:
        c.name = body.name.strip()
    if body.slug is not None:
        new_slug = body.slug.strip()
        other = db.scalar(select(Category).where(Category.slug == new_slug, Category.id != c.id))
        if other is not None:
            raise HTTPException(status_code=400, detail="Slug already in use")
        c.slug = new_slug
    db.commit()
    db.refresh(c)
    return CategoryOut(id=str(c.id), name=c.name, slug=c.slug)


@router.delete("/categories/{category_id}", response_model=OkOut)
def admin_delete_category(category_id: str, db: Session = Depends(get_db)):
    c = db.get(Category, category_id)
    if not c:
        raise HTTPException(status_code=404, detail="Category not found")
    cnt = db.scalar(select(func.count()).select_from(Part).where(Part.category_id == c.id)) or 0
    if cnt > 0:
        raise HTTPException(status_code=400, detail="Category has parts; reassign or delete parts first")
    db.delete(c)
    db.commit()
    return OkOut()


def _category_id_for_slug(db: Session, slug: str | None) -> uuid.UUID | None:
    if not slug or not str(slug).strip():
        return None
    s = str(slug).strip()
    cat = db.scalar(select(Category).where(Category.slug == s))
    if not cat:
        raise HTTPException(status_code=400, detail=f"Unknown category slug: {s}")
    return cat.id


def _resolve_category_slug(
    db: Session, cat_slug: str | None
) -> tuple[uuid.UUID | None, str | None]:
    if not cat_slug or not str(cat_slug).strip():
        return None, None
    s = str(cat_slug).strip()
    cat = db.scalar(select(Category).where(Category.slug == s))
    if not cat:
        return None, f"Unknown category_slug: {s}"
    return cat.id, None


def _part_to_out(part: Part, db: Session) -> PartOut:
    cat_slug = None
    if part.category_id:
        cat = db.get(Category, part.category_id)
        cat_slug = cat.slug if cat else None
    return PartOut(
        id=str(part.id),
        part_number=part.part_number,
        name=part.name,
        description=part.description,
        category=cat_slug,
        stock_quantity=part.stock_quantity,
        status=part.status,
        price=float(part.price) if part.price is not None else None,
    )


@router.get("/parts", response_model=list[PartOut])
def admin_list_parts(
    search: str | None = Query(default=None),
    category: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
    db: Session = Depends(get_db),
):
    q = select(Part)
    if category:
        q = q.join(Category, Part.category_id == Category.id).where(Category.slug == category)
    if search and search.strip():
        like = f"%{search.strip()}%"
        q = q.where(
            or_(
                Part.part_number.ilike(like),
                Part.name.ilike(like),
            )
        )
    q = q.order_by(Part.part_number.asc()).limit(limit)
    rows = db.execute(q).scalars().all()
    return [_part_to_out(p, db) for p in rows]


@router.post("/parts", response_model=PartOut)
async def admin_create_part(body: PartCreate, db: Session = Depends(get_db)):
    if db.scalar(select(Part).where(Part.part_number == body.part_number.strip())) is not None:
        raise HTTPException(status_code=400, detail="Part number already exists")
    cid = _category_id_for_slug(db, body.category_slug)
    p = Part(
        id=uuid.uuid4(),
        part_number=body.part_number.strip(),
        name=body.name.strip(),
        description=body.description,
        category_id=cid,
        stock_quantity=body.stock_quantity,
        price=Decimal(str(body.price)) if body.price is not None else None,
        status=body.status,
    )
    db.add(p)
    db.flush()
    refresh_part_search_vector(db, p)
    await notify_subscribers_if_became_available(db, p, was_available=False)
    db.commit()
    db.refresh(p)
    return _part_to_out(p, db)


@router.patch("/parts/{part_id}", response_model=PartOut)
async def admin_update_part(part_id: str, body: PartUpdate, db: Session = Depends(get_db)):
    p = db.get(Part, part_id)
    if not p:
        raise HTTPException(status_code=404, detail="Part not found")
    was_available = is_part_available(p)
    if body.part_number is not None:
        pn = body.part_number.strip()
        other = db.scalar(select(Part).where(Part.part_number == pn, Part.id != p.id))
        if other is not None:
            raise HTTPException(status_code=400, detail="Part number already in use")
        p.part_number = pn
    if body.name is not None:
        p.name = body.name.strip()
    if body.description is not None:
        p.description = body.description
    if body.category_slug is not None:
        cs = body.category_slug.strip()
        if not cs:
            p.category_id = None
        else:
            p.category_id = _category_id_for_slug(db, cs)
    if body.stock_quantity is not None:
        p.stock_quantity = body.stock_quantity
    if body.price is not None:
        p.price = Decimal(str(body.price))
    if body.status is not None:
        p.status = body.status
    db.flush()
    refresh_part_search_vector(db, p)
    await notify_subscribers_if_became_available(db, p, was_available=was_available)
    db.commit()
    db.refresh(p)
    return _part_to_out(p, db)


@router.delete("/parts/{part_id}", response_model=OkOut)
def admin_delete_part(part_id: str, db: Session = Depends(get_db)):
    p = db.get(Part, part_id)
    if not p:
        raise HTTPException(status_code=404, detail="Part not found")
    db.delete(p)
    db.commit()
    return OkOut()


EXPECTED_HEADERS = (
    "part_number",
    "name",
    "description",
    "category_slug",
    "stock_quantity",
    "price",
    "status",
)


def _normalize_header(h: str | None) -> str:
    if h is None:
        return ""
    return h.strip().lower().replace(" ", "_")


def _parse_rows_from_csv(raw: bytes) -> list[dict[str, str]]:
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV has no header row")
    fieldmap = {_normalize_header(f): f for f in reader.fieldnames if f}
    rows: list[dict[str, str]] = []
    for line in reader:
        row: dict[str, str] = {}
        for key in EXPECTED_HEADERS:
            src = fieldmap.get(key)
            row[key] = (line.get(src, "") if src else "") or ""
        rows.append(row)
    return rows


def _parse_rows_from_xlsx(raw: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel file is empty")
    headers = [_normalize_header(str(c) if c is not None else "") for c in header_row]
    idx = {h: i for i, h in enumerate(headers) if h}

    def cell(r: tuple[Any, ...], key: str) -> str:
        i = idx.get(key)
        if i is None or i >= len(r):
            return ""
        v = r[i]
        if v is None:
            return ""
        return str(v).strip()

    out: list[dict[str, str]] = []
    for r in rows_iter:
        if all((cell(r, k) == "" for k in EXPECTED_HEADERS)):
            continue
        out.append({k: cell(r, k) for k in EXPECTED_HEADERS})
    return out


@router.post("/parts/import", response_model=ImportResult)
async def admin_import_parts(
    file: UploadFile = File(...),
    dry_run: bool = Query(default=False),
    db: Session = Depends(get_db),
):
    raw = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".csv"):
        try:
            data_rows = _parse_rows_from_csv(raw)
        except UnicodeDecodeError as e:
            raise HTTPException(status_code=400, detail=f"CSV must be UTF-8: {e}") from e
    elif name.endswith(".xlsx"):
        data_rows = _parse_rows_from_xlsx(raw)
    else:
        raise HTTPException(status_code=400, detail="Upload a .csv or .xlsx file")

    result = ImportResult()
    for i, row in enumerate(data_rows, start=2):
        part_number = (row.get("part_number") or "").strip()
        name_part = (row.get("name") or "").strip()
        if not part_number or not name_part:
            result.errors.append(ImportRowError(row=i, message="part_number and name are required"))
            continue
        desc = (row.get("description") or "").strip() or None
        cat_slug = (row.get("category_slug") or "").strip() or None
        try:
            stock = int(float((row.get("stock_quantity") or "0").strip() or "0"))
        except ValueError:
            result.errors.append(ImportRowError(row=i, message="stock_quantity must be a number"))
            continue
        price_raw = (row.get("price") or "").strip()
        price: Decimal | None = None
        if price_raw:
            try:
                price = Decimal(price_raw.replace("$", "").replace(",", ""))
            except Exception:
                result.errors.append(ImportRowError(row=i, message="invalid price"))
                continue
        status = (row.get("status") or "in_stock").strip() or "in_stock"
        if len(status) > 24:
            result.errors.append(ImportRowError(row=i, message="status too long"))
            continue

        cid, cat_err = _resolve_category_slug(db, cat_slug)
        if cat_err:
            result.errors.append(ImportRowError(row=i, message=cat_err))
            continue

        existing = db.scalar(select(Part).where(Part.part_number == part_number))
        if dry_run:
            if existing:
                result.updated += 1
            else:
                result.created += 1
            continue

        if existing:
            was_available = is_part_available(existing)
            existing.name = name_part
            existing.description = desc
            existing.category_id = cid
            existing.stock_quantity = stock
            existing.price = price
            existing.status = status
            db.flush()
            refresh_part_search_vector(db, existing)
            await notify_subscribers_if_became_available(db, existing, was_available=was_available)
            db.commit()
            result.updated += 1
        else:
            p = Part(
                id=uuid.uuid4(),
                part_number=part_number,
                name=name_part,
                description=desc,
                category_id=cid,
                stock_quantity=stock,
                price=price,
                status=status,
            )
            db.add(p)
            db.flush()
            refresh_part_search_vector(db, p)
            await notify_subscribers_if_became_available(db, p, was_available=False)
            db.commit()
            result.created += 1

    return result


@router.post("/outreach/send", response_model=OutreachSendOut)
async def admin_outreach_send(
    body: OutreachSendIn,
):
    from app.email import send_bulk_emails

    recipients = [str(e) for e in body.recipients]
    n = await send_bulk_emails(recipients, body.subject, body.body)
    return OutreachSendOut(sent=n)


@router.get("/inventory-alerts", response_model=list[dict])
def admin_list_inventory_alerts(db: Session = Depends(get_db)):
    from app.models import InventoryAlertSubscription

    rows = (
        db.execute(
            select(InventoryAlertSubscription, Part)
            .join(Part, InventoryAlertSubscription.part_id == Part.id)
            .order_by(InventoryAlertSubscription.created_at.desc())
        )
        .all()
    )
    out = []
    for sub, part in rows:
        out.append(
            {
                "id": str(sub.id),
                "email": sub.email,
                "part_number": part.part_number,
                "part_name": part.name,
                "active": sub.active,
                "created_at": sub.created_at.isoformat(),
                "last_notified_at": sub.last_notified_at.isoformat()
                if sub.last_notified_at
                else None,
            }
        )
    return out
